import tkinter as tk
from tkinter import filedialog, ttk
import torch
import torch.nn.functional as F
from torchvision import transforms
from PIL import Image, ImageTk, ImageFilter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import os
from model import creer_modele

# ══════════════════════════════════════════
#  CONFIGURATION
# ══════════════════════════════════════════
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "models", "meilleur_modele.pth")
DEVICE     = torch.device("cpu")
CLASSES    = ["Femelle", "Male"]
SEUIL      = 0.70

BG       = "#1e1e2e"
CARD     = "#2a2a3e"
ACCENT   = "#7c3aed"
TEXT     = "#ffffff"
SUBTEXT  = "#a0a0b0"
SUCCESS  = "#22c55e"
WARNING  = "#f59e0b"
DANGER   = "#ef4444"
MALE_C   = "#60a5fa"
FEMALE_C = "#f472b6"

# ══════════════════════════════════════════
#  MODELE
# ══════════════════════════════════════════
transform = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.Grayscale(num_output_channels=3),
    transforms.ToTensor(),
    transforms.Normalize([0.485, 0.456, 0.406],
                         [0.229, 0.224, 0.225])
])

def charger_modele():
    modele = creer_modele()
    modele.load_state_dict(torch.load(MODEL_PATH, map_location=DEVICE))
    modele.eval()
    return modele

def predire(modele, chemin):
    # Validation morphologique préalable pour écarter les images hors-sujet (ex: serpents, paysages)
    try:
        img_temp = Image.open(chemin).convert("L")
        img_temp = img_temp.resize((224, 224))
        arr_temp = np.array(img_temp, dtype=np.float32)
        h, w     = arr_temp.shape
        marge    = int(0.2 * min(h, w))
        zone     = arr_temp[marge:h-marge, marge:w-marge]

        # 1. Luminosité (ratio_clair) : les rémiges sont très sombres
        pixels_clairs = (zone > 128).sum()
        ratio_clair   = (pixels_clairs / zone.size) * 100

        # 2. Contraste : écart-type de l'intensité
        contraste = float(zone.std())

        # 3. Densité de contours
        img_pil = Image.fromarray(zone.astype(np.uint8))
        contours = np.array(img_pil.filter(ImageFilter.FIND_EDGES), dtype=np.float32)
        densite_contours = (contours > 30).sum() / contours.size * 100

        # Seuils de validation morphologique basés sur le jeu de données réel
        if ratio_clair > 5.0 or contraste < 12.0 or contraste > 50.0 or densite_contours < 1.5 or densite_contours > 15.0:
            return "Non conforme", 0.0, 0.0, 0.0
    except:
        return "Non conforme", 0.0, 0.0, 0.0

    image  = Image.open(chemin).convert("RGB")
    tensor = transform(image).unsqueeze(0)
    with torch.no_grad():
        sortie = modele(tensor)
        proba  = F.softmax(sortie, dim=1)
        conf, pred = torch.max(proba, 1)
    prob_female = proba[0][0].item() * 100
    prob_male   = proba[0][1].item() * 100
    confiance   = conf.item() * 100
    if confiance < SEUIL * 100:
        classe = "Indetermine"
    else:
        classe = CLASSES[pred.item()]
    return classe, confiance, prob_male, prob_female

def analyser_remiges(chemin):
    img      = Image.open(chemin).convert("L")
    img      = img.resize((224, 224))
    arr      = np.array(img, dtype=np.float32)
    h, w     = arr.shape
    marge    = int(0.2 * min(h, w))
    zone     = arr[marge:h-marge, marge:w-marge]

    pixels_clairs    = (zone > 128).sum()
    ratio_clair      = (pixels_clairs / zone.size) * 100
    contraste        = float(zone.std())
    img_pil          = Image.fromarray(zone.astype(np.uint8))
    contours         = np.array(
        img_pil.filter(ImageFilter.FIND_EDGES), dtype=np.float32)
    densite_contours = (contours > 30).sum() / contours.size * 100
    zone_haute       = arr[:h//2, marge:w-marge]
    zone_basse       = arr[h//2:, marge:w-marge]
    asymetrie        = abs(zone_haute.mean() - zone_basse.mean())
    score_femelle    = (
        (ratio_clair      / 100) * 0.30 +
        (contraste        / 128) * 0.35 +
        (densite_contours / 100) * 0.25 +
        (asymetrie        / 128) * 0.10
    ) * 100

    return {
        "ratio_clair"      : round(ratio_clair,      2),
        "contraste"        : round(contraste,         2),
        "densite_contours" : round(densite_contours,  2),
        "asymetrie"        : round(asymetrie,         2),
        "score_femelle"    : round(score_femelle,     2),
        "score_male"       : round(100 - score_femelle, 2),
    }

# ══════════════════════════════════════════
#  APPLICATION
# ══════════════════════════════════════════
class App:
    def __init__(self, root):
        self.root      = root
        self.modele    = None
        self.resultats = []
        self.chemins   = {}
        self.setup_window()
        self.build_ui()
        self.charger_modele_async()

    def setup_window(self):
        self.root.title("Sexage Poussins par Remiges — ResNet18")
        self.root.geometry("1100x900")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)
        x = (self.root.winfo_screenwidth()  // 2) - 550
        y = (self.root.winfo_screenheight() // 2) - 450
        self.root.geometry(f"1100x900+{x}+{y}")

    def build_ui(self):
        # TITRE
        titre = tk.Frame(self.root, bg=ACCENT, pady=12)
        titre.pack(fill="x")
        tk.Label(titre,
                 text="Sexage Automatique des Poussins par Remiges",
                 font=("Segoe UI", 18, "bold"),
                 bg=ACCENT, fg=TEXT).pack()
        tk.Label(titre,
                 text="ResNet18 — Precision : 96.69%  |  AUC-ROC : 0.974",
                 font=("Segoe UI", 10),
                 bg=ACCENT, fg="#d4b8ff").pack()

        # STATUS
        sf = tk.Frame(self.root, bg=BG, pady=5)
        sf.pack(fill="x", padx=20)
        self.status = tk.Label(sf, text="Chargement du modele...",
                               font=("Segoe UI", 10),
                               bg=BG, fg=WARNING)
        self.status.pack()

        # ZONE PRINCIPALE
        main = tk.Frame(self.root, bg=BG)
        main.pack(fill="both", expand=True, padx=15, pady=5)

        # Colonne gauche — image
        left = tk.Frame(main, bg=CARD, width=420)
        left.pack(side="left", fill="y", padx=(0, 8))
        left.pack_propagate(False)
        tk.Label(left, text="Image analysee",
                 font=("Segoe UI", 11, "bold"),
                 bg=CARD, fg=TEXT, pady=8).pack()
        self.img_label = tk.Label(left, bg=CARD,
                                  text="Aucune image",
                                  font=("Segoe UI", 11),
                                  fg=SUBTEXT)
        self.img_label.pack(expand=True)
        self.nom_label = tk.Label(left, text="",
                                  font=("Segoe UI", 8),
                                  bg=CARD, fg=SUBTEXT,
                                  wraplength=400)
        self.nom_label.pack(pady=4)

        # Colonne milieu — resultats
        mid = tk.Frame(main, bg=CARD, width=300)
        mid.pack(side="left", fill="y", padx=(0, 8))
        mid.pack_propagate(False)
        tk.Label(mid, text="Resultat",
                 font=("Segoe UI", 11, "bold"),
                 bg=CARD, fg=TEXT, pady=8).pack()
        self.resultat_label = tk.Label(mid, text="—",
                                       font=("Segoe UI", 32, "bold"),
                                       bg=CARD, fg=SUBTEXT)
        self.resultat_label.pack(pady=8)
        self.conf_label = tk.Label(mid, text="",
                                   font=("Segoe UI", 12),
                                   bg=CARD, fg=SUBTEXT)
        self.conf_label.pack()
        self.verdict_label = tk.Label(mid, text="",
                                      font=("Segoe UI", 10),
                                      bg=CARD, fg=SUBTEXT,
                                      wraplength=280)
        self.verdict_label.pack(pady=4)

        bframe = tk.Frame(mid, bg=CARD, pady=10)
        bframe.pack(fill="x", padx=15)
        tk.Label(bframe, text="Male",
                 font=("Segoe UI", 10, "bold"),
                 bg=CARD, fg=MALE_C).pack(anchor="w")
        self.barre_male_var = tk.DoubleVar()
        ttk.Progressbar(bframe, variable=self.barre_male_var,
                        maximum=100, mode="determinate").pack(
                            fill="x", pady=2)
        self.pct_male = tk.Label(bframe, text="0.0%",
                                  font=("Segoe UI", 9),
                                  bg=CARD, fg=MALE_C)
        self.pct_male.pack(anchor="e")
        tk.Label(bframe, text="Femelle",
                 font=("Segoe UI", 10, "bold"),
                 bg=CARD, fg=FEMALE_C).pack(anchor="w", pady=(8,0))
        self.barre_female_var = tk.DoubleVar()
        ttk.Progressbar(bframe, variable=self.barre_female_var,
                        maximum=100, mode="determinate").pack(
                            fill="x", pady=2)
        self.pct_female = tk.Label(bframe, text="0.0%",
                                    font=("Segoe UI", 9),
                                    bg=CARD, fg=FEMALE_C)
        self.pct_female.pack(anchor="e")

        # Colonne droite — tableau
        right = tk.Frame(main, bg=CARD)
        right.pack(side="right", fill="both", expand=True)
        tk.Label(right, text="Fichiers analyses",
                 font=("Segoe UI", 11, "bold"),
                 bg=CARD, fg=TEXT, pady=8).pack()
        self.compteur = tk.Label(right, text="",
                                  font=("Segoe UI", 9),
                                  bg=CARD, fg=SUBTEXT)
        self.compteur.pack()

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background=CARD, foreground=TEXT,
                        rowheight=26, fieldbackground=CARD,
                        font=("Segoe UI", 9))
        style.configure("Treeview.Heading",
                        background="#3a3a5e", foreground=TEXT,
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", ACCENT)])

        cols = ("fichier", "sexe", "confiance", "male", "femelle")
        self.tableau = ttk.Treeview(right, columns=cols,
                                    show="headings",
                                    selectmode="browse")
        self.tableau.heading("fichier",   text="Fichier")
        self.tableau.heading("sexe",      text="Sexe")
        self.tableau.heading("confiance", text="Confiance")
        self.tableau.heading("male",      text="% Male")
        self.tableau.heading("femelle",   text="% Femelle")
        self.tableau.column("fichier",   width=160, anchor="w")
        self.tableau.column("sexe",      width=90,  anchor="center")
        self.tableau.column("confiance", width=80,  anchor="center")
        self.tableau.column("male",      width=70,  anchor="center")
        self.tableau.column("femelle",   width=70,  anchor="center")
        self.tableau.tag_configure("male",
                                   background="#1a2a3d",
                                   foreground=MALE_C)
        self.tableau.tag_configure("femelle",
                                   background="#2d1f3d",
                                   foreground=FEMALE_C)
        self.tableau.tag_configure("indet",
                                   background="#2a2a1f",
                                   foreground=WARNING)
        sb = ttk.Scrollbar(right, orient="vertical",
                           command=self.tableau.yview)
        self.tableau.configure(yscrollcommand=sb.set)
        self.tableau.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.tableau.bind("<<TreeviewSelect>>", self.on_select)

        # BOUTONS
        bf = tk.Frame(self.root, bg=BG, pady=8)
        bf.pack()
        self.btn_img = tk.Button(
            bf, text="Image",
            font=("Segoe UI", 11, "bold"),
            bg=ACCENT, fg=TEXT, relief="flat",
            padx=18, pady=8, cursor="hand2",
            command=self.choisir_image, state="disabled")
        self.btn_img.pack(side="left", padx=6)
        self.btn_dir = tk.Button(
            bf, text="Dossier",
            font=("Segoe UI", 11, "bold"),
            bg="#1d4ed8", fg=TEXT, relief="flat",
            padx=18, pady=8, cursor="hand2",
            command=self.choisir_dossier, state="disabled")
        self.btn_dir.pack(side="left", padx=6)
        self.btn_hist = tk.Button(
            bf, text="Histogrammes",
            font=("Segoe UI", 11, "bold"),
            bg="#065f46", fg=TEXT, relief="flat",
            padx=18, pady=8, cursor="hand2",
            command=self.afficher_histogrammes, state="disabled")
        self.btn_hist.pack(side="left", padx=6)
        self.btn_csv = tk.Button(
            bf, text="Exporter CSV",
            font=("Segoe UI", 11, "bold"),
            bg="#7c2d12", fg=TEXT, relief="flat",
            padx=18, pady=8, cursor="hand2",
            command=self.exporter_csv, state="disabled")
        self.btn_csv.pack(side="left", padx=6)

        # FOOTER
        footer = tk.Frame(self.root, bg="#13131f", pady=4)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer,
                 text="ResNet18 | Transfer Learning | PyTorch 2.12 | CPU | Seuil 70%",
                 font=("Segoe UI", 8),
                 bg="#13131f", fg=SUBTEXT).pack()

    def charger_modele_async(self):
        self.root.after(100, self._charger)

    def _charger(self):
        try:
            self.modele = charger_modele()
            self.status.config(
                text="Modele pret — Precision : 96.69%  |  AUC : 0.974",
                fg=SUCCESS)
            self.btn_img.config(state="normal")
            self.btn_dir.config(state="normal")
        except Exception as e:
            self.status.config(text=f"Erreur : {e}", fg=DANGER)

    def choisir_image(self):
        chemin = filedialog.askopenfilename(
            filetypes=[("Images",
                        "*.png *.jpg *.jpeg *.bmp *.tiff")])
        if not chemin:
            return
        self.afficher_image(chemin)
        classe, conf, pm, pf = predire(self.modele, chemin)
        self.maj_resultat(classe, conf, pm, pf)
        self.tableau.delete(*self.tableau.get_children())
        self.chemins   = {}
        self.resultats = []
        self.ajouter_ligne(chemin, classe, conf, pm, pf)
        self.resultats.append({
            "fichier": os.path.basename(chemin),
            "sexe": classe, "confiance": conf,
            "male": pm, "femelle": pf})
        self.maj_compteur()
        self.btn_hist.config(state="normal")
        self.btn_csv.config(state="normal")

    def choisir_dossier(self):
        dossier = filedialog.askdirectory()
        if not dossier:
            return
        exts   = (".png", ".jpg", ".jpeg", ".bmp", ".tiff")
        images = sorted([f for f in os.listdir(dossier)
                         if f.lower().endswith(exts)])
        if not images:
            self.verdict_label.config(
                text="Aucune image trouvee", fg=DANGER)
            return
        self.tableau.delete(*self.tableau.get_children())
        self.chemins   = {}
        self.resultats = []
        nb_m = nb_f = nb_i = 0
        self.status.config(
            text=f"Analyse en cours... 0 / {len(images)}",
            fg=WARNING)
        self.root.update()
        for i, nom in enumerate(images):
            chemin = os.path.join(dossier, nom)
            try:
                classe, conf, pm, pf = predire(
                    self.modele, chemin)
            except:
                classe, conf, pm, pf = "Indetermine", 0, 0, 0
            self.ajouter_ligne(chemin, classe, conf, pm, pf)
            self.resultats.append({
                "fichier": nom, "sexe": classe,
                "confiance": conf,
                "male": pm, "femelle": pf})
            if classe == "Male":
                nb_m += 1
            elif classe == "Femelle":
                nb_f += 1
            else:
                nb_i += 1
            self.status.config(
                text=f"Analyse... {i+1}/{len(images)}")
            self.root.update()
        self.status.config(
            text=f"Analyse terminee — {len(images)} images",
            fg=SUCCESS)
        self.resultat_label.config(text="Dossier", fg=ACCENT)
        self.conf_label.config(
            text=f"{len(images)} images analysees")
        self.verdict_label.config(
            text=f"Males: {nb_m}  Femelles: {nb_f}  "
                 f"Indetermines: {nb_i}",
            fg=SUCCESS)
        self.maj_compteur()
        self.btn_hist.config(state="normal")
        self.btn_csv.config(state="normal")

    def ajouter_ligne(self, chemin, classe, conf, pm, pf):
        nom = os.path.basename(chemin)
        if classe == "Male":
            tag = "male"
        elif classe == "Femelle":
            tag = "femelle"
        else:
            tag = "indet"
        iid = self.tableau.insert("", "end",
            values=(nom, classe, f"{conf:.1f}%",
                    f"{pm:.1f}%", f"{pf:.1f}%"),
            tags=(tag,))
        self.chemins[iid] = chemin

    def on_select(self, event):
        sel = self.tableau.selection()
        if not sel:
            return
        chemin = self.chemins.get(sel[0])
        if chemin and os.path.exists(chemin):
            self.afficher_image(chemin)
            classe, conf, pm, pf = predire(self.modele, chemin)
            self.maj_resultat(classe, conf, pm, pf)

    def afficher_image(self, chemin):
        img = Image.open(chemin).convert("RGB")
        img.thumbnail((390, 370))
        photo = ImageTk.PhotoImage(img)
        self.img_label.config(image=photo, text="")
        self.img_label.image = photo
        self.img_label.bind("<Button-1>",
            lambda e: self.ouvrir_plein_ecran(chemin))
        self.nom_label.config(
            text=os.path.basename(chemin) +
                 "  (clic = plein ecran)")

    def ouvrir_plein_ecran(self, chemin):
        win = tk.Toplevel(self.root)
        win.title(os.path.basename(chemin))
        win.configure(bg="black")
        win.attributes("-fullscreen", True)
        win.bind("<Escape>",   lambda e: win.destroy())
        win.bind("<Button-1>", lambda e: win.destroy())
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        img = Image.open(chemin).convert("RGB")
        img.thumbnail((sw, sh), Image.LANCZOS)
        photo = ImageTk.PhotoImage(img)
        lbl = tk.Label(win, image=photo, bg="black")
        lbl.image = photo
        lbl.pack(expand=True)
        tk.Label(win,
                 text=os.path.basename(chemin) +
                      "  |  Echap pour fermer",
                 font=("Segoe UI", 11),
                 bg="black", fg="white").pack(pady=8)

    def maj_resultat(self, classe, conf, pm, pf):
        self.barre_male_var.set(pm)
        self.barre_female_var.set(pf)
        self.pct_male.config(text=f"{pm:.1f}%")
        self.pct_female.config(text=f"{pf:.1f}%")
        self.conf_label.config(text=f"Confiance : {conf:.1f}%")
        if classe == "Male":
            self.resultat_label.config(text="Male",     fg=MALE_C)
            self.verdict_label.config(
                text=f"C'est un male ! ({conf:.1f}%)",
                fg=SUCCESS)
        elif classe == "Femelle":
            self.resultat_label.config(text="Femelle",  fg=FEMALE_C)
            self.verdict_label.config(
                text=f"C'est une femelle ! ({conf:.1f}%)",
                fg=SUCCESS)
        elif classe == "Non conforme":
            self.resultat_label.config(text="Indet.",   fg=WARNING)
            self.verdict_label.config(
                text="Image non conforme (pas un poussin/rémige)",
                fg=WARNING)
        else:
            self.resultat_label.config(text="Indet.",   fg=WARNING)
            self.verdict_label.config(
                text=f"Confiance insuffisante ({conf:.1f}% < {SEUIL*100:.0f}%)",
                fg=WARNING)

    def maj_compteur(self):
        total = len(self.tableau.get_children())
        nb_m  = sum(1 for i in self.tableau.get_children()
                    if "male"    in self.tableau.item(i)["tags"])
        nb_f  = sum(1 for i in self.tableau.get_children()
                    if "femelle" in self.tableau.item(i)["tags"])
        nb_i  = total - nb_m - nb_f
        self.compteur.config(
            text=f"Total: {total}  |  Males: {nb_m}  "
                 f"Femelles: {nb_f}  Indet: {nb_i}")

    def afficher_histogrammes(self):
        if not self.resultats:
            return

        males    = [r for r in self.resultats
                    if r["sexe"] == "Male"]
        femelles = [r for r in self.resultats
                    if r["sexe"] == "Femelle"]
        indets   = [r for r in self.resultats
                    if r["sexe"] not in ["Male", "Femelle"]]

        win = tk.Toplevel(self.root)
        win.title("Histogrammes — Analyse complete")
        win.geometry("1100x750")
        win.configure(bg=BG)

        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=10, pady=10)

        # ── ONGLET 1 — Stats generales ────────
        tab1 = tk.Frame(nb, bg=BG)
        nb.add(tab1, text="  Statistiques generales  ")

        fig1, axes1 = plt.subplots(2, 2, figsize=(12, 7))
        fig1.patch.set_facecolor("#1e1e2e")
        plt.style.use("dark_background")

        labels = ["Males", "Femelles", "Indetermines"]
        vals   = [len(males), len(femelles), len(indets)]
        colors = [MALE_C, FEMALE_C, WARNING]

        ax = axes1[0][0]
        bars = ax.bar(labels, vals, color=colors,
                      edgecolor="white", lw=0.5)
        ax.set_title("Repartition par Sexe",
                     fontsize=12, fontweight="bold", color=TEXT)
        ax.set_facecolor(CARD)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width()/2,
                    v + 0.2, str(v), ha="center",
                    fontweight="bold", color=TEXT, fontsize=12)

        ax = axes1[0][1]
        conf_m = [r["confiance"] for r in males]
        conf_f = [r["confiance"] for r in femelles]
        if conf_m:
            ax.hist(conf_m, bins=15, alpha=0.7, color=MALE_C,
                    label=f"Males ({len(conf_m)})",
                    edgecolor="white", lw=0.5)
        if conf_f:
            ax.hist(conf_f, bins=15, alpha=0.7, color=FEMALE_C,
                    label=f"Femelles ({len(conf_f)})",
                    edgecolor="white", lw=0.5)
        ax.axvline(70, color="yellow", linestyle="--",
                   lw=1.5, label="Seuil 70%")
        ax.set_title("Distribution des Confiances",
                     fontsize=12, fontweight="bold", color=TEXT)
        ax.set_xlabel("Confiance (%)", color=SUBTEXT)
        ax.set_ylabel("Nombre", color=SUBTEXT)
        ax.set_facecolor(CARD)
        ax.legend(fontsize=9)

        ax = axes1[1][0]
        moy_m = sum(conf_m)/len(conf_m) if conf_m else 0
        moy_f = sum(conf_f)/len(conf_f) if conf_f else 0
        bars3 = ax.bar(["Males", "Femelles"], [moy_m, moy_f],
                       color=[MALE_C, FEMALE_C],
                       edgecolor="white", lw=0.5)
        ax.set_ylim(0, 100)
        ax.axhline(70, color="yellow", linestyle="--",
                   lw=1.5, label="Seuil 70%")
        ax.set_title("Confiance Moyenne par Classe",
                     fontsize=12, fontweight="bold", color=TEXT)
        ax.set_ylabel("Confiance (%)", color=SUBTEXT)
        ax.set_facecolor(CARD)
        ax.legend(fontsize=9)
        for bar, v in zip(bars3, [moy_m, moy_f]):
            ax.text(bar.get_x() + bar.get_width()/2,
                    v + 0.5, f"{v:.1f}%", ha="center",
                    fontweight="bold", color=TEXT, fontsize=11)

        ax = axes1[1][1]
        pvals  = [v for v in vals if v > 0]
        plbls  = [l for l, v in zip(labels, vals) if v > 0]
        pcolrs = [c for c, v in zip(colors, vals) if v > 0]
        if pvals:
            ax.pie(pvals, labels=plbls, colors=pcolrs,
                   autopct="%1.1f%%", startangle=90,
                   textprops={"color": TEXT, "fontsize": 11},
                   wedgeprops={"edgecolor": BG, "linewidth": 2})
        ax.set_title("Repartition en %",
                     fontsize=12, fontweight="bold", color=TEXT)
        ax.set_facecolor(CARD)

        plt.suptitle("Statistiques Generales du Lot",
                     fontsize=14, fontweight="bold", color=TEXT)
        plt.tight_layout()
        FigureCanvasTkAgg(fig1, master=tab1).get_tk_widget()\
            .pack(fill="both", expand=True)

        # ── ONGLET 2 — Analyse pixels remiges ─
        tab2 = tk.Frame(nb, bg=BG)
        nb.add(tab2, text="  Analyse Pixels Remiges  ")

        self.status.config(
            text="Analyse des remiges...", fg=WARNING)
        win.update()

        analyses_m, analyses_f = [], []
        for r in males:
            c = next((v for iid, v in self.chemins.items()
                      if os.path.basename(v) == r["fichier"]),
                     None)
            if c:
                try:
                    analyses_m.append(analyser_remiges(c))
                except:
                    pass
        for r in femelles:
            c = next((v for iid, v in self.chemins.items()
                      if os.path.basename(v) == r["fichier"]),
                     None)
            if c:
                try:
                    analyses_f.append(analyser_remiges(c))
                except:
                    pass

        self.status.config(
            text="Modele pret — Precision : 96.69%", fg=SUCCESS)

        fig2, axes2 = plt.subplots(2, 3, figsize=(13, 7))
        fig2.patch.set_facecolor("#1e1e2e")

        metriques_info = [
            ("ratio_clair",
             "Ratio Pixels Clairs (%)",
             "pixels > 128 / total x 100"),
            ("contraste",
             "Contraste Local",
             "Ecart-type intensite pixels"),
            ("densite_contours",
             "Densite de Contours (%)",
             "Bords detectes / total x 100"),
            ("asymetrie",
             "Asymetrie Haut/Bas",
             "|moy_haute - moy_basse|"),
            ("score_femelle",
             "Score Morpho Femelle (%)",
             "0.30*Clair+0.35*Contrast"
             "+0.25*Contours+0.10*Asym"),
            ("score_male",
             "Score Morpho Male (%)",
             "100 - Score_Femelle"),
        ]

        for idx, (cle, titre, formule) in \
                enumerate(metriques_info):
            ax  = axes2[idx // 3][idx % 3]
            vm  = [a[cle] for a in analyses_m] if analyses_m \
                else [0]
            vf  = [a[cle] for a in analyses_f] if analyses_f \
                else [0]
            moy_mv = sum(vm)/len(vm) if vm else 0
            moy_fv = sum(vf)/len(vf) if vf else 0
            bars = ax.bar([0, 1], [moy_mv, moy_fv],
                          color=[MALE_C, FEMALE_C],
                          edgecolor="white", lw=0.5, width=0.5)
            ax.set_xticks([0, 1])
            ax.set_xticklabels(["Male", "Femelle"],
                               fontsize=10, color=TEXT)
            ax.set_title(titre, fontsize=10,
                         fontweight="bold", color=TEXT)
            ax.set_facecolor(CARD)
            ax.tick_params(colors=TEXT)
            for bar, v in zip(bars, [moy_mv, moy_fv]):
                ax.text(bar.get_x() + bar.get_width()/2,
                        v + 0.3, f"{v:.1f}",
                        ha="center", fontweight="bold",
                        color=TEXT, fontsize=10)
            ax.text(0.5, -0.22, formule, ha="center",
                    transform=ax.transAxes, fontsize=7,
                    color=SUBTEXT, style="italic")

        plt.suptitle(
            "Analyse Morphologique des Remiges par Pixels",
            fontsize=12, fontweight="bold", color=TEXT)
        plt.tight_layout()
        FigureCanvasTkAgg(fig2, master=tab2).get_tk_widget()\
            .pack(fill="both", expand=True)

        tk.Label(tab2,
                 text="Formule : Score_Femelle(%) = "
                      "0.30 x Pixels_Clairs + "
                      "0.35 x Contraste + "
                      "0.25 x Densite_Contours + "
                      "0.10 x Asymetrie  |  "
                      "Score_Male = 100 - Score_Femelle",
                 font=("Segoe UI", 9), bg=BG,
                 fg=SUBTEXT, justify="center").pack(pady=5)

        # Bouton export
        def exporter_tout():
            dossier = filedialog.askdirectory(
                title="Choisir dossier export")
            if not dossier:
                return
            fig1.savefig(os.path.join(dossier,
                         "1_stats_generales.png"),
                         dpi=150, bbox_inches="tight")
            fig2.savefig(os.path.join(dossier,
                         "2_analyse_remiges.png"),
                         dpi=150, bbox_inches="tight")
            self.status.config(
                text="Graphiques exportes !", fg=SUCCESS)

        tk.Button(win,
                  text="Exporter tous les graphiques (PNG)",
                  font=("Segoe UI", 11, "bold"),
                  bg=ACCENT, fg=TEXT, relief="flat",
                  padx=20, pady=8, cursor="hand2",
                  command=exporter_tout).pack(pady=8)

    def exporter_csv(self):
        chemin = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not chemin:
            return

        import openpyxl
        from openpyxl.styles import (PatternFill, Font,
                                      Alignment, Border, Side)
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.image import Image as XLImage
        import tempfile

        self.status.config(
            text="Generation du fichier Excel...", fg=WARNING)
        self.root.update()

        wb  = openpyxl.Workbook()

        # ══════════════════════════════════════════
        # FEUILLE 1 — DONNEES
        # ══════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Donnees"

        # Style en-tete
        fill_header = PatternFill("solid", fgColor="7C3AED")
        font_header = Font(bold=True, color="FFFFFF", size=11)
        font_titre  = Font(bold=True, color="FFFFFF", size=14)
        fill_male   = PatternFill("solid", fgColor="1A2A3D")
        fill_female = PatternFill("solid", fgColor="2D1F3D")
        fill_indet  = PatternFill("solid", fgColor="2A2A1F")
        font_male   = Font(color="60A5FA", size=10)
        font_female = Font(color="F472B6", size=10)
        font_indet  = Font(color="F59E0B", size=10)
        border_side = Side(style="thin", color="CCCCCC")
        border      = Border(
            left=border_side, right=border_side,
            top=border_side,  bottom=border_side)
        centre      = Alignment(horizontal="center",
                                vertical="center")

        # Titre
        ws1.merge_cells("A1:K1")
        ws1["A1"] = "RAPPORT DE SEXAGE AUTOMATIQUE DES POUSSINS"
        ws1["A1"].font      = font_titre
        ws1["A1"].fill      = PatternFill("solid", fgColor="1E1E2E")
        ws1["A1"].alignment = centre

        ws1.merge_cells("A2:K2")
        ws1["A2"] = ("ResNet18 | Precision : 96.69% | "
                     "AUC-ROC : 0.974 | Seuil : 70%")
        ws1["A2"].font      = Font(italic=True, color="A0A0B0",
                                   size=10)
        ws1["A2"].fill      = PatternFill("solid", fgColor="2A2A3E")
        ws1["A2"].alignment = centre

        # En-tetes colonnes
        headers = [
            "Fichier", "Sexe", "Confiance",
            "% Male", "% Femelle",
            "Pixels Clairs (%)", "Contraste",
            "Densite Contours (%)", "Asymetrie",
            "Score Male (%)", "Score Femelle (%)"
        ]
        for col, h in enumerate(headers, 1):
            cell            = ws1.cell(row=4, column=col, value=h)
            cell.font       = font_header
            cell.fill       = fill_header
            cell.alignment  = centre
            cell.border     = border

        # Largeurs colonnes
        largeurs = [25, 12, 12, 10, 12, 18, 12, 20, 12, 15, 18]
        for col, larg in enumerate(largeurs, 1):
            ws1.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = larg

        # Donnees
        row = 5
        for iid in self.tableau.get_children():
            vals       = self.tableau.item(iid)["values"]
            chemin_img = self.chemins.get(iid, "")
            tags       = self.tableau.item(iid)["tags"]

            # Metriques remiges
            if chemin_img and os.path.exists(chemin_img):
                try:
                    m = analyser_remiges(chemin_img)
                    r_vals = [
                        m["ratio_clair"],
                        m["contraste"],
                        m["densite_contours"],
                        m["asymetrie"],
                        m["score_male"],
                        m["score_femelle"]
                    ]
                except:
                    r_vals = ["N/A"] * 6
            else:
                r_vals = ["N/A"] * 6

            all_vals = list(vals) + r_vals

            # Choisir style selon sexe
            if "male" in tags:
                fill_row = fill_male
                font_row = font_male
            elif "femelle" in tags:
                fill_row = fill_female
                font_row = font_female
            else:
                fill_row = fill_indet
                font_row = font_indet

            for col, val in enumerate(all_vals, 1):
                cell            = ws1.cell(row=row,
                                           column=col,
                                           value=val)
                cell.fill       = fill_row
                cell.font       = font_row
                cell.alignment  = centre
                cell.border     = border
            row += 1

        # Ligne resume
        ws1.cell(row=row+1, column=1,
                 value="RESUME").font = Font(bold=True,
                                             color="FFFFFF")
        items = self.tableau.get_children()
        nb_m  = sum(1 for i in items
                    if "male"    in self.tableau.item(i)["tags"])
        nb_f  = sum(1 for i in items
                    if "femelle" in self.tableau.item(i)["tags"])
        nb_i  = len(items) - nb_m - nb_f

        resume = [
            ("Total images",    len(items)),
            ("Males",           nb_m),
            ("Femelles",        nb_f),
            ("Indetermines",    nb_i),
            ("Precision modele","96.69%"),
            ("AUC-ROC",         "0.974"),
        ]
        for r_idx, (label, val) in enumerate(resume):
            ws1.cell(row=row+1, column=r_idx*2+1,
                     value=label).font = Font(bold=True,
                                              color="A0A0B0",
                                              size=9)
            ws1.cell(row=row+2, column=r_idx*2+1,
                     value=val).font   = Font(bold=True,
                                              color="FFFFFF",
                                              size=11)

        # ══════════════════════════════════════════
        # FEUILLE 2 — GRAPHIQUES
        # ══════════════════════════════════════════
        ws2 = wb.create_sheet("Graphiques")
        ws2["A1"] = "GRAPHIQUES D'ANALYSE"
        ws2["A1"].font = Font(bold=True, color="7C3AED", size=14)

        # Donnees pour graphiques
        ws2["A3"]  = "Classe"
        ws2["B3"]  = "Nombre"
        ws2["A4"]  = "Males"
        ws2["B4"]  = nb_m
        ws2["A5"]  = "Femelles"
        ws2["B5"]  = nb_f
        ws2["A6"]  = "Indetermines"
        ws2["B6"]  = nb_i

        # Graphique barres — Repartition
        chart1 = BarChart()
        chart1.type        = "col"
        chart1.title       = "Repartition Males / Femelles"
        chart1.y_axis.title = "Nombre"
        chart1.x_axis.title = "Classe"
        chart1.style       = 10
        chart1.width       = 15
        chart1.height      = 12

        data1 = Reference(ws2, min_col=2, min_row=3,
                          max_row=6)
        cats1 = Reference(ws2, min_col=1, min_row=4,
                          max_row=6)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        ws2.add_chart(chart1, "D3")

        # Graphique camembert — Pourcentages
        chart2      = PieChart()
        chart2.title = "Repartition en Pourcentage"
        chart2.style = 10
        chart2.width  = 15
        chart2.height = 12

        data2 = Reference(ws2, min_col=2, min_row=3,
                          max_row=6)
        cats2 = Reference(ws2, min_col=1, min_row=4,
                          max_row=6)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws2.add_chart(chart2, "D22")

        # Donnees confiances pour graphique
        ws2["A10"] = "Classe"
        ws2["B10"] = "Conf. Moyenne (%)"
        conf_m = [r["confiance"] for r in self.resultats
                  if r["sexe"] == "Male"]
        conf_f = [r["confiance"] for r in self.resultats
                  if r["sexe"] == "Femelle"]
        ws2["A11"] = "Males"
        ws2["B11"] = round(sum(conf_m)/len(conf_m), 2) \
                     if conf_m else 0
        ws2["A12"] = "Femelles"
        ws2["B12"] = round(sum(conf_f)/len(conf_f), 2) \
                     if conf_f else 0

        chart3 = BarChart()
        chart3.type         = "col"
        chart3.title        = "Confiance Moyenne par Classe"
        chart3.y_axis.title = "Confiance (%)"
        chart3.y_axis.scaling.min = 0
        chart3.y_axis.scaling.max = 100
        chart3.style        = 10
        chart3.width        = 15
        chart3.height       = 12
        data3 = Reference(ws2, min_col=2, min_row=10,
                          max_row=12)
        cats3 = Reference(ws2, min_col=1, min_row=11,
                          max_row=12)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)
        ws2.add_chart(chart3, "M3")

        # Sauvegarder
        wb.save(chemin)
        self.status.config(
            text=f"Excel exporte : {os.path.basename(chemin)}",
            fg=SUCCESS)

# ══════════════════════════════════════════
#  LANCEMENT
# ══════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    app  = App(root)
    root.mainloop()